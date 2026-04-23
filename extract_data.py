#!/usr/bin/env python3
"""
Régénère data/foods.json, data/foods.js, data/clinical.json, data/clinical.js
depuis document_maitre_tyramine_nardil_v4.xlsx (golden source).

Usage : python3 extract_data.py

Ne modifier jamais les fichiers data/ manuellement.
Pour mettre à jour le dataset, modifier le xlsx puis relancer ce script.
"""
import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("Dépendance manquante : pip install openpyxl")

ROOT = Path(__file__).parent
XLSX = ROOT / "document_maitre_tyramine_nardil_v4.xlsx"
DATA = ROOT / "data"
DATA.mkdir(exist_ok=True)


def clean(v):
    if v is None:
        return None
    if isinstance(v, bool):
        return v
    if isinstance(v, float):
        return int(v) if v == int(v) else round(v, 4)
    return v


def extract_foods(wb):
    ws = wb["Aliments_consolide"]
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[2]
    foods = []
    for row in rows[3:]:
        if row[0] is None:
            continue
        entry = {headers[i]: clean(row[i]) for i in range(len(headers))}
        foods.append(entry)
    return foods


def extract_clinical(wb):
    # Seuils_cliniques
    ws = wb["Seuils_cliniques"]
    rows = [r for r in ws.iter_rows(values_only=True) if any(x not in (None, "") for x in r)]
    seuils = []
    for r in rows[2:]:
        if r[0] and r[1]:
            seuils.append({"parametre": r[0], "valeur": r[1], "source": r[2], "implication": r[3]})

    # Facteurs_algorithme
    ws = wb["Facteurs_algorithme"]
    rows = [r for r in ws.iter_rows(values_only=True) if any(x not in (None, "") for x in r)]
    facteurs = []
    for r in rows[3:]:
        if r[0] and r[0] not in ("Garde-fou",) and r[2] is not None:
            facteurs.append({"facteur": r[0], "condition": r[1], "coeff": r[2], "justification": r[3]})

    # Regles_alerte
    ws = wb["Regles_alerte"]
    rows = [r for r in ws.iter_rows(values_only=True) if any(x not in (None, "") for x in r)]
    regles = []
    for r in rows[2:]:
        if r[0] in ("NOIR", "ROUGE", "ORANGE", "VERT LIMITÉ"):
            regles.append({"niveau": r[0], "regle": r[1], "sens_clinique": r[2], "affichage_patient": r[3]})

    # Wording_patient
    ws = wb["Wording_patient"]
    rows = [r for r in ws.iter_rows(values_only=True) if any(x not in (None, "") for x in r)]
    wording = {}
    sections = [
        "Mode d'emploi patient",
        "Bloc urgence (visible sans clic)",
        "Disclaimer fort",
        "Projet de clause droit et juridiction",
    ]
    current = None
    for r in rows:
        if r[0] in sections:
            current = r[0]
        elif current and r[0]:
            wording[current] = r[0]
            current = None

    # Signaux_majeurs
    ws = wb["Signaux_majeurs"]
    rows = [r for r in ws.iter_rows(values_only=True) if any(x not in (None, "") for x in r)]
    signaux = []
    for r in rows[2:]:
        if r[0] is not None and r[1]:
            signaux.append({"num": r[0], "signal": r[1], "implication": r[2]})

    # Sources
    ws = wb["Sources"]
    rows = [r for r in ws.iter_rows(values_only=True) if any(x not in (None, "") for x in r)]
    sources = []
    for r in rows[2:]:
        if r[1] and r[0] not in ("Type",):
            sources.append({"type": r[0], "reference": r[1], "url": r[2], "usage": r[3]})

    return {
        "seuils_cliniques": seuils,
        "formule": "score = tyramine_mg_portion_typ × f_dose × f_anciennete × f_incertitude",
        "facteurs": facteurs,
        "regles_alerte": regles,
        "wording": wording,
        "signaux_majeurs": signaux,
        "sources": sources,
    }


def write_js(path, var_name, data, header_lines):
    with open(path, "w", encoding="utf-8") as f:
        for line in header_lines:
            f.write(f"// {line}\n")
        f.write(f"\nwindow.{var_name} = ")
        f.write(json.dumps(data, ensure_ascii=False, indent=2))
        f.write(";\n")


def main():
    if not XLSX.exists():
        sys.exit(f"Fichier introuvable : {XLSX}")

    print(f"Chargement de {XLSX.name}...")
    wb = openpyxl.load_workbook(XLSX, data_only=True)

    foods = extract_foods(wb)
    clinical = extract_clinical(wb)

    (DATA / "foods.json").write_text(json.dumps(foods, ensure_ascii=False, indent=2), encoding="utf-8")
    write_js(
        DATA / "foods.js",
        "FOODS_DATA",
        foods,
        [
            "Golden source — 338 entrées extraites de document_maitre_tyramine_nardil_v4.xlsx",
            "Schéma : id | nom_fr | nom_ru_si_pertinent | cuisine_region | procede | unite_base",
            "       | mg100_low | mg100_typ | mg100_high | portion_standard",
            "       | mgportion_low | mgportion_typ | mgportion_high",
            "       | confiance (A/B/C/D) | note | source_url | alias_menu_restaurant",
            "       | artisanal_sensitive | label_restricted | variabilite",
            "       | note_complementaire_rapport | categorie_libelle_fda | seuil_alerte_par_portion",
            "ND (données manquantes confiance D) = null ; le verrou seuil_alerte_par_portion reste NOIR",
            "Ne pas modifier manuellement ; régénérer depuis le xlsx via extract_data.py",
        ],
    )

    (DATA / "clinical.json").write_text(json.dumps(clinical, ensure_ascii=False, indent=2), encoding="utf-8")
    write_js(
        DATA / "clinical.js",
        "CLINICAL",
        clinical,
        [
            "Données cliniques — seuils, facteurs, règles, signaux, wording, sources",
            "Source : document_maitre_tyramine_nardil_v4.xlsx",
            "Ne pas modifier manuellement ; régénérer via extract_data.py",
        ],
    )

    print(f"✓ {len(foods)} aliments extraits")
    print(f"✓ data/foods.json  ({(DATA/'foods.json').stat().st_size:,} octets)")
    print(f"✓ data/foods.js    ({(DATA/'foods.js').stat().st_size:,} octets)")
    print(f"✓ data/clinical.json ({(DATA/'clinical.json').stat().st_size:,} octets)")
    print(f"✓ data/clinical.js   ({(DATA/'clinical.js').stat().st_size:,} octets)")


if __name__ == "__main__":
    main()
